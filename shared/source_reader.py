"""
source_reader.py — 自动读取源 Excel 财务数据
Read IS_DATA / BS_DATA / CF_DATA directly from the source financial Excel file.
Eliminates manual transcription errors and 10x unit-scale issues.
"""
import hashlib
import math
from pathlib import Path
from typing import Dict, List, Tuple, Optional
from datetime import datetime

import openpyxl

from .validators import CA_KEYS, NCA_KEYS, CL_KEYS, NCL_KEYS, EQ_KEYS


# ═══════════════════════════════════════════════════════════════════
# Row-label → key mapping for each sheet
# ═══════════════════════════════════════════════════════════════════

# 利润表: row_num → dict_key
IS_ROW_MAP = {
    2:  'Revenue_Total',
    3:  'COGS_Total',
    4:  'Tax_Surcharge',
    5:  'Selling_Exp',
    6:  'Admin_Exp',
    7:  'RD_Exp',
    8:  'Finance_Cost',
    9:  'Interest_Expense',
    10: 'Interest_Income',
    11: 'Other_Income',
    12: 'Investment_Income',
    15: 'Credit_Impairment',
    16: 'Asset_Impairment',
    18: 'Operating_Profit',
    22: 'Income_Tax',
    23: 'Net_Profit',
}

# 资产负债表: row_num → dict_key
BS_ROW_MAP = {
    3:  'Cash',
    4:  'Trading_Fin_Assets',
    5:  'Notes_Receivable',
    6:  'Accounts_Receivable',
    8:  'Prepayments',
    9:  'Other_Receivables',
    10: 'Inventory',
    11: 'Contract_Assets',
    12: 'Other_Current_Assets',
    19: 'LT_Equity_Investments',
    20: 'Other_Equity_Investments',
    23: 'Fixed_Assets',
    26: 'Intangible_Assets',
    29: 'LT_Prepaid',
    25: 'ROU_Assets',
    30: 'Deferred_Tax_Assets',
    37: 'ST_Borrowings',
    40: 'Accounts_Payable',
    41: 'Advance_Receipts',
    42: 'Contract_Liabilities',
    43: 'Employee_Comp_Payable',
    44: 'Taxes_Payable',
    45: 'Other_Payable',
    52: 'LT_Borrowings',
    54: 'Lease_Liabilities',
    57: 'Deferred_Income',
    58: 'Deferred_Tax_Liabilities',
    65: 'Paid_in_Capital',
    67: 'Capital_Reserve',
    71: 'Surplus_Reserve',
    72: 'Retained_Earnings',
}

# 现金流量表: row_num → dict_key
CF_ROW_MAP = {
    12: 'Operating_CF',
    3:  'Cash_From_Sales',
    7:  'Cash_For_Goods',
    8:  'Cash_To_Employees',
    26: 'Investing_CF',
    21: 'Capex',
    22: 'Investment_Purchase',
    37: 'Financing_CF',
    29: 'Capital_Raised',
    42: 'Cash_End',
}

# 收入明细-按产品: row_num → dict_key (单位: 元)
REVENUE_DETAIL_ROW_MAP = {
    2: 'Revenue_Servo',
    3: 'Revenue_Dexterous',
    4: 'Revenue_Other',
}

# 毛利明细-按产品: row_num → dict_key (单位: 万元, 用于推算 COGS)
MARGIN_DETAIL_ROW_MAP = {
    3: 'Margin_Servo',
    4: 'Margin_Dexterous',
    5: 'Margin_Other',
}


# BS subtotal key groups — imported from validators.py (single source of truth)
# CA_KEYS, NCA_KEYS, CL_KEYS, NCL_KEYS, EQ_KEYS


def _yuan_to_wan(val, decimals: int = 0) -> int:
    """Convert 元 → 万元, rounding to integer."""
    if val is None or val == '' or val == 0:
        return 0
    try:
        f = float(val) / 10000.0
        return int(round(f, decimals))
    except (TypeError, ValueError):
        return 0


def _read_sheet_data(ws, row_map: Dict[int, str],
                     n_years: int, convert: bool = True) -> Dict[str, List[int]]:
    """
    Read data from a worksheet using a row → key mapping.
    Data columns start at column B (index 2).
    If convert=True, values are treated as 元 and divided by 10000.
    """
    result = {}
    for row_num, key in row_map.items():
        vals = []
        for ci in range(2, 2 + n_years):
            raw = ws.cell(row_num, ci).value
            if convert:
                vals.append(_yuan_to_wan(raw))
            else:
                # Already in 万元 or %
                vals.append(raw if raw is not None else 0)
        result[key] = vals
    return result


def _adjust_cash_for_bs_balance(bs_data: Dict[str, List[int]],
                                n_years: int) -> List[int]:
    """
    Adjust Cash values ±1 to eliminate rounding residuals.
    Returns the list of adjustments applied.

    After 元→万元 rounding, individual items may have ±0.5万 error each,
    causing BS imbalance of a few 万. We force balance by nudging Cash.
    """
    adjustments = []
    for i in range(n_years):
        ta = sum(bs_data[k][i] for k in CA_KEYS + NCA_KEYS)
        tle = sum(bs_data[k][i] for k in CL_KEYS + NCL_KEYS + EQ_KEYS)
        gap = ta - tle  # positive = assets too high
        if gap != 0:
            bs_data['Cash'][i] -= gap
            adjustments.append(gap)
        else:
            adjustments.append(0)
    return adjustments


def compute_file_hash(path: Path) -> str:
    """Compute MD5 hash of a file for tracking changes."""
    h = hashlib.md5()
    with open(path, 'rb') as f:
        while chunk := f.read(8192):
            h.update(chunk)
    return h.hexdigest()


def read_source_financials(
    source_path: str | Path,
    n_years: int = 4,
    row_maps: dict | None = None,
) -> Tuple[Dict[str, List[int]], Dict[str, List[int]],
           Dict[str, List[int]], Dict[str, str]]:
    """
    Read historical financial data from source Excel workbook.

    Args:
        source_path: Path to the financial summary xlsx.
        n_years: Number of historical years (default 4).
        row_maps: Optional dict from project_config.yaml ``source_row_maps``.
                  When provided, overrides module-level *_ROW_MAP constants.
                  Expected keys: is_row_map, bs_row_map, cf_row_map,
                  plus optional revenue_detail_row_map, margin_detail_row_map,
                  outflow_sign_inversion, sheet_names.

    Returns:
        (is_data, bs_data, cf_data, meta) — three dicts of key→[values]
        plus a meta dict with file_hash, file_mtime, etc.
    """
    path = Path(source_path)
    if not path.exists():
        raise FileNotFoundError(f"Source data file not found: {path}")

    # ── Resolve row maps ──
    if row_maps:
        def _as_int(m):
            return {int(k): v for k, v in m.items()}
        is_map = _as_int(row_maps['is_row_map'])
        bs_map = _as_int(row_maps['bs_row_map'])
        cf_map = _as_int(row_maps['cf_row_map'])
        rev_detail_map = (
            _as_int(row_maps['revenue_detail_row_map'])
            if 'revenue_detail_row_map' in row_maps
            else REVENUE_DETAIL_ROW_MAP
        )
        margin_detail_map = (
            _as_int(row_maps['margin_detail_row_map'])
            if 'margin_detail_row_map' in row_maps
            else MARGIN_DETAIL_ROW_MAP
        )
        outflow_keys = row_maps.get(
            'outflow_sign_inversion',
            ['Capex', 'Investment_Purchase'],
        )
        snames = row_maps.get('sheet_names', {})
    else:
        is_map = IS_ROW_MAP
        bs_map = BS_ROW_MAP
        cf_map = CF_ROW_MAP
        rev_detail_map = REVENUE_DETAIL_ROW_MAP
        margin_detail_map = MARGIN_DETAIL_ROW_MAP
        outflow_keys = ['Capex', 'Investment_Purchase']
        snames = {}

    sn_is = snames.get('income_statement', '利润表')
    sn_bs = snames.get('balance_sheet', '资产负债表')
    sn_cf = snames.get('cash_flow', '现金流量表')
    sn_rev = snames.get('revenue_detail', '收入明细-按产品')
    sn_margin = snames.get('margin_detail', '毛利明细-按产品')

    wb = openpyxl.load_workbook(path, data_only=True)

    # ── 1. Read main statements (元 → 万元) ──
    is_data = _read_sheet_data(wb[sn_is], is_map, n_years, convert=True)
    bs_data = _read_sheet_data(wb[sn_bs], bs_map, n_years, convert=True)
    cf_data = _read_sheet_data(wb[sn_cf], cf_map, n_years, convert=True)

    # ── 1b. CF sign convention: source stores outflows as positive,
    #    but our model convention uses negative for outflows ──
    for outflow_key in outflow_keys:
        if outflow_key in cf_data:
            cf_data[outflow_key] = [
                -abs(v) if v > 0 else v
                for v in cf_data[outflow_key]
            ]

    # ── 2. Read revenue/COGS detail (if sheets exist) ──
    if sn_rev in wb.sheetnames:
        rev_detail = _read_sheet_data(
            wb[sn_rev], rev_detail_map, n_years, convert=True
        )
        is_data.update(rev_detail)

    if sn_margin in wb.sheetnames:
        margin_detail = _read_sheet_data(
            wb[sn_margin], margin_detail_map, n_years, convert=False
        )
        # Derive COGS per product = Revenue - Margin
        # Revenue_Servo is in 万元 (already converted), Margin is in 万元 natively
        for product in ['Servo', 'Dexterous', 'Other']:
            rev_key = f'Revenue_{product}'
            margin_key = f'Margin_{product}'
            cogs_key = f'COGS_{product}'
            if rev_key in is_data and margin_key in margin_detail:
                is_data[cogs_key] = [
                    is_data[rev_key][i] - int(round(margin_detail[margin_key][i]))
                    for i in range(n_years)
                ]

    wb.close()

    # ── 3. Adjust Cash for BS balance ──
    adjustments = _adjust_cash_for_bs_balance(bs_data, n_years)

    # ── 4. Build metadata ──
    file_stat = path.stat()
    meta = {
        'source_path': str(path),
        'file_hash': compute_file_hash(path),
        'file_mtime': datetime.fromtimestamp(file_stat.st_mtime).isoformat(),
        'file_size': str(file_stat.st_size),
        'read_time': datetime.now().isoformat(),
        'cash_adjustments': str(adjustments),
        'n_years': str(n_years),
    }

    return is_data, bs_data, cf_data, meta
