"""
style_utils.py — 统一样式定义
所有 build 脚本共用的颜色常量、字体/填充/对齐/边框工厂函数、数字格式。
"""
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ═══════════════════════════════════════════════════════════════
# COLOR PALETTE (hex, no '#')
# ═══════════════════════════════════════════════════════════════
C_TITLE   = '1F4E79'   # 深蓝 — 标题栏
C_BLK     = '2E75B6'   # 蓝 — Block header
C_SUBBLK  = '5B9BD5'   # 浅蓝 — 子 Block header
C_COLHDR  = 'D6E4F0'   # 极浅蓝 — 列标题
C_CORE    = 'EBF5FB'   # 核心可比行
C_REF     = 'FDFEFE'   # 参考可比行
C_STAT    = 'D5E8D4'   # 统计行 (绿底)
C_SUBJ    = 'FFF2CC'   # 主体锚点 (黄底)
C_VAL     = 'E2EFDA'   # 估值推算 (浅绿)
C_BRIDGE  = 'FCE4D6'   # 桥接行 (橙底)
C_GRAY    = 'F2F2F2'   # 灰色
C_WHITE   = 'FFFFFF'
C_NOTE    = 'EBF5FB'   # 备注/结论
C_DCF     = 'D5E8D4'   # DCF行 (绿)
C_PE      = 'EBF5FB'   # P/E行 (浅蓝)
C_PS      = 'FFF2CC'   # P/S行 (黄)
C_PB      = 'FCE4D6'   # P/B行 (橙)
C_WGT     = 'E2EFDA'   # 加权行
C_CUR     = 'F9F9F9'   # 当前融资行
C_HIGHLIGHT = 'FFF2CC' # 高亮

# ═══════════════════════════════════════════════════════════════
# BORDERS
# ═══════════════════════════════════════════════════════════════
THIN_SIDE    = Side(style='thin',   color='BFBFBF')
THIN_SIDE_D9 = Side(style='thin',   color='D9D9D9')
MED_SIDE     = Side(style='medium', color='808080')
DBL_SIDE     = Side(style='double', color='000000')

THIN_BORDER  = Border(bottom=THIN_SIDE_D9)
DOUBLE_BORDER = Border(top=THIN_SIDE_D9, bottom=DBL_SIDE)

# ═══════════════════════════════════════════════════════════════
# NUMBER FORMATS
# ═══════════════════════════════════════════════════════════════
FMT_TEXT = '@'
FMT_N0   = '#,##0'
FMT_N1   = '#,##0.0'
FMT_N2   = '#,##0.00'
FMT_PCT  = '0.0%'
FMT_PCT2 = '0.00%'

# ═══════════════════════════════════════════════════════════════
# FACTORY FUNCTIONS
# ═══════════════════════════════════════════════════════════════
def fill(hex_: str) -> PatternFill:
    """Create a solid fill from a hex color string (with or without '#')."""
    return PatternFill('solid', fgColor=hex_.lstrip('#'))


def fnt(bold=False, size=10, color='000000', italic=False,
        name='Calibri') -> Font:
    """Create a Font object."""
    return Font(bold=bold, size=size, color=color, italic=italic, name=name)


def fnt_yahei(bold=False, size=10, color='000000', italic=False) -> Font:
    """Create a Microsoft YaHei Font (used in block1/block2 scripts)."""
    return Font(bold=bold, size=size, color=color, italic=italic,
                name='Microsoft YaHei')


def aln(h='left', v='center', wrap=False) -> Alignment:
    """Create an Alignment object."""
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def bdr(all_thin=False, top=False, bot=False, lft=False, rgt=False,
        med_bot=False, med_top=False) -> Border:
    """Create a Border object with optional thin/medium sides."""
    kw = {}
    if all_thin:
        kw = dict(top=THIN_SIDE, bottom=THIN_SIDE,
                  left=THIN_SIDE, right=THIN_SIDE)
    if top or med_top:
        kw['top'] = MED_SIDE
    if bot or med_bot:
        kw['bottom'] = MED_SIDE
    if lft:
        kw['left'] = MED_SIDE
    if rgt:
        kw['right'] = MED_SIDE
    return Border(**kw)


# ═══════════════════════════════════════════════════════════════
# PRE-BUILT STYLE OBJECTS (for block1 / block2 scripts)
# ═══════════════════════════════════════════════════════════════
HEADER_FILL  = PatternFill(start_color='1F4E79', end_color='1F4E79',
                           fill_type='solid')
SECTION_FILL = PatternFill(start_color='D6E4F0', end_color='D6E4F0',
                           fill_type='solid')

HDR_FONT_W    = fnt_yahei(bold=True,  size=10, color='FFFFFF')
SECTION_FONT  = fnt_yahei(bold=True,  size=10, color='1F4E79')
BOLD_FONT     = fnt_yahei(bold=True,  size=10)
DATA_FONT     = fnt_yahei(size=10)
INLINE_FONT   = fnt_yahei(size=9, italic=True, color='1F4E79')
CHECK_FONT    = fnt_yahei(bold=True,  size=10, color='006100')
TITLE_FONT    = fnt_yahei(bold=True,  size=14, color='1F4E79')


# ═══════════════════════════════════════════════════════════════
# THREE-STATEMENT STYLE PACK (IS/BS/CF/Returns/Cross_Check/DCF)
# ── 命名沿用 session_b-g 约定，方便一次 import 替换 ──
# ═══════════════════════════════════════════════════════════════

# Fonts
FNT_TITLE   = TITLE_FONT                                        # 14pt 深蓝粗
FNT_HDR     = HDR_FONT_W                                        # 10pt 白字粗
FNT_SEC     = SECTION_FONT                                      # 10pt 深蓝粗
FNT_DATA    = DATA_FONT                                         # 10pt 正常
FNT_BOLD    = BOLD_FONT                                         # 10pt 粗
FNT_INPUT   = Font(name='Calibri', size=10, color='000080')     # 蓝色输入字
FNT_GREEN   = Font(name='Calibri', size=10, color='008000')     # 绿色公式字
FNT_NOTE    = INLINE_FONT                                       # 9pt 斜体深蓝
FNT_CHECK   = CHECK_FONT                                        # 粗绿色
FNT_RESULT  = fnt_yahei(bold=True, size=12, color='1F4E79')     # 12pt 结果

# Fills
FILL_HDR    = HEADER_FILL                                       # 深蓝
FILL_SEC    = SECTION_FILL                                      # 浅蓝
FILL_INPUT  = PatternFill('solid', fgColor='EBF5FB')            # 极浅蓝
FILL_WHITE  = PatternFill('solid', fgColor='FFFFFF')
FILL_RESULT = PatternFill('solid', fgColor='E2EFDA')            # 浅绿
FILL_GREEN  = PatternFill('solid', fgColor='C6EFCE')            # 绿色
FILL_BASE   = PatternFill('solid', fgColor='C6EFCE')            # Scenario 基
FILL_BEAR   = PatternFill('solid', fgColor='FCD5B4')            # Scenario 熊
FILL_BULL   = PatternFill('solid', fgColor='D9EAD3')            # Scenario 牛

# Alignment shortcuts
ALN_L = Alignment(horizontal='left',   vertical='center')
ALN_C = Alignment(horizontal='center', vertical='center')
ALN_R = Alignment(horizontal='right',  vertical='center')
