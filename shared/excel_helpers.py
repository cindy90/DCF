"""
excel_helpers.py — 通用 Excel 写入辅助函数
Block header、column header、data cell 写入函数。
支持 Comps/ValSummary 风格 (write_blk_hdr / write_col_hdrs / write_cell)
和 block1/block2 风格 (make_header_row)。
"""
from openpyxl.utils import get_column_letter
from .style_utils import (
    fill, fnt, fnt_yahei, aln, bdr,
    C_BLK, C_COLHDR, C_WHITE, FMT_TEXT,
    HEADER_FILL, HDR_FONT_W, SECTION_FILL, SECTION_FONT,
    BOLD_FONT, DATA_FONT, INLINE_FONT, THIN_BORDER,
    # Three-statement style imports (used by create_statement_sheet)
    FNT_TITLE, FNT_HDR, FILL_HDR, ALN_C,
)


# ═══════════════════════════════════════════════════════════════
# THREE-STATEMENT SHEET HELPERS (IS / BS / CF / Returns / etc.)
# ═══════════════════════════════════════════════════════════════

def create_statement_sheet(wb, name: str, title: str, year_labels: list,
                           position: int | None = 0, col_a_width: int = 32,
                           note_col_width: int = 42):
    """
    Create (or replace) a financial-statement sheet with standard formatting.

    - Deletes any existing sheet with the same name
    - Creates a new sheet at ``position`` (None = append at end)
    - Merges row 1 for the title, applies FNT_TITLE
    - Writes column headers in row 3: ['项目'] + year_labels + ['备注']
    - Sets column widths and disables gridlines

    Returns the new worksheet object.
    """
    if name in wb.sheetnames:
        del wb[name]
    if position is not None:
        ws = wb.create_sheet(name, position)
    else:
        ws = wb.create_sheet(name)
    n_year_cols = len(year_labels)
    end_col = get_column_letter(1 + n_year_cols + 1)  # A + years + note

    # ── Title row ──
    ws.merge_cells(f'A1:{end_col}1')
    ws['A1'] = title
    ws['A1'].font = FNT_TITLE

    # ── Column headers (row 3) ──
    headers = ['项目'] + year_labels + ['备注']
    for ci, h in enumerate(headers):
        c = ws.cell(3, ci + 1, h)
        c.font = FNT_HDR
        c.fill = FILL_HDR
        c.alignment = ALN_C

    # ── Column widths ──
    ws.column_dimensions['A'].width = col_a_width
    for i in range(n_year_cols):
        ws.column_dimensions[get_column_letter(i + 2)].width = 14
    ws.column_dimensions[get_column_letter(n_year_cols + 2)].width = note_col_width

    ws.sheet_view.showGridLines = False
    return ws


# ═══════════════════════════════════════════════════════════════
# COMPS / VALUATION SUMMARY STYLE HELPERS
# ═══════════════════════════════════════════════════════════════

def write_blk_hdr(ws, row, text, merge_end='S', color=C_BLK, h=18):
    """Write a colored section header bar, merged from B to merge_end."""
    ws.row_dimensions[row].height = h
    ws.merge_cells(f'B{row}:{merge_end}{row}')
    c = ws[f'B{row}']
    c.value     = text
    c.font      = fnt(bold=True, size=11, color='FFFFFF')
    c.fill      = fill(color)
    c.alignment = aln('left')
    c.border    = bdr(all_thin=True)


def write_col_hdrs(ws, row, hdrs, col_start='B'):
    """Write a row of column headers with standard blue styling."""
    ws.row_dimensions[row].height = 28
    ci = ord(col_start) - ord('A')
    for i, h in enumerate(hdrs):
        col = get_column_letter(ci + i + 1)
        c = ws[f'{col}{row}']
        c.value     = h
        c.font      = fnt(bold=True, size=9, color='1F4E79')
        c.fill      = fill(C_COLHDR)
        c.alignment = aln('center', wrap=True)
        c.border    = bdr(all_thin=True)


def write_cell(ws, row, col, val, fmt=FMT_TEXT, bold=False, bg=C_WHITE,
               color='000000', h_align='center', wrap=False, italic=False):
    """Write a single formatted cell in the Comps/ValSummary style."""
    c = ws[f'{col}{row}']
    c.value         = val
    c.font          = fnt(bold=bold, size=9, color=color, italic=italic)
    c.fill          = fill(bg)
    c.alignment     = aln(h_align, wrap=wrap)
    c.number_format = fmt
    c.border        = bdr(all_thin=True)
    return c


# ═══════════════════════════════════════════════════════════════
# BLOCK1 / BLOCK2 STYLE HELPERS
# ═══════════════════════════════════════════════════════════════

def make_header_row(ws, title_text, year_labels, note_col='J'):
    """
    Write row 1 title + row 3 column headers for Returns/Cross_Check/Summary.
    year_labels: list like ['FY2023','FY2024','FY2025','FY2026E',...]
    """
    # Merge range depends on number of year columns
    n_cols = len(year_labels)
    end_col = get_column_letter(1 + n_cols + 1)  # A + years + note
    ws.merge_cells(f'A1:{end_col}1')
    ws['A1'] = title_text
    ws['A1'].font = fnt_yahei(bold=True, size=12, color='1F4E79')

    headers = ['项目'] + year_labels + ['备注']
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(3, ci, h)
        cell.font      = HDR_FONT_W
        cell.fill       = HEADER_FILL
        cell.alignment  = aln('center')
        cell.border     = THIN_BORDER

    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 40
    for i, _ in enumerate(year_labels):
        ws.column_dimensions[get_column_letter(2 + i)].width = 16
    ws.column_dimensions[get_column_letter(2 + n_cols)].width = 42


def add_section_row(ws, row, label, n_data_cols=9):
    """Write a section header row with blue background (block1/block2 style)."""
    ws.cell(row, 1, label).font = SECTION_FONT
    ws.cell(row, 1).fill = SECTION_FILL
    for ci in range(2, n_data_cols + 2):
        ws.cell(row, ci).fill = SECTION_FILL


def add_data_row(ws, row, label, formula_template, all_cols,
                 fmt='#,##0.00', note='', bold=False):
    """
    Write a data row with formula for each year column.
    formula_template: str with '{col}' placeholder, e.g. '=IS!{col}17'
    all_cols: list of Excel column letters, e.g. ['B','C','D','E','F','G','H','I']
    """
    font = BOLD_FONT if bold else DATA_FONT
    ws.cell(row, 1, label).font = font
    for ci, col_l in enumerate(all_cols):
        cell = ws.cell(row, ci + 2)
        cell.value = formula_template.format(col=col_l)
        cell.font = font
        cell.number_format = fmt
        cell.border = THIN_BORDER
    if note:
        note_col = len(all_cols) + 2
        ws.cell(row, note_col).value = note
        ws.cell(row, note_col).font = INLINE_FONT


def setup_gridlines(wb, exclude=('_State',)):
    """Disable gridlines on all sheets except those in `exclude`."""
    for name in wb.sheetnames:
        if name not in exclude:
            wb[name].sheet_view.showGridLines = False
